"""
Sistema de Auto-Mejora Programado para Bruce W
Se ejecuta automáticamente cada viernes a las 9:00 AM
Requiere autorización manual del usuario antes de aplicar cambios
"""

import os
import json
import schedule
import time
from datetime import datetime, timedelta
from auto_mejora_bruce import AutoMejoraBruce
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generar_excel_analisis(analisis, mejoras_seleccionadas=None):
    """
    Genera un Excel completo con el análisis semanal

    Args:
        analisis: Diccionario con el análisis completo
        mejoras_seleccionadas: Lista de números de mejoras aprobadas (opcional)

    Returns:
        str: Ruta del archivo Excel generado
    """
    fecha_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_archivo = f"analisis_bruce_{fecha_str}.xlsx"

    # Extraer datos
    stats = analisis['stats']
    analisis_gpt = analisis['analisis']
    fecha_analisis = analisis['fecha']

    # ========================================
    # SHEET 1: RESUMEN EJECUTIVO
    # ========================================

    resumen_data = {
        'MÉTRICA': [
            'Período Analizado',
            'Fecha de Análisis',
            '',
            'Total de Llamadas',
            'Llamadas Aprobadas',
            'Llamadas Negadas',
            'Tasa de Conversión (%)',
            'WhatsApps Capturados',
            '',
            'Nivel de Interés Promedio',
            'Estado de Ánimo Predominante',
        ],
        'VALOR': [
            'Últimos 7 días',
            fecha_analisis,
            '',
            stats['total_llamadas'],
            stats['aprobados'],
            stats['total_llamadas'] - stats['aprobados'],
            f"{stats['tasa_conversion']:.2f}%",
            stats['whatsapps_capturados'],
            '',
            stats['interes_promedio'],
            stats['animo_predominante'],
        ]
    }

    df_resumen = pd.DataFrame(resumen_data)

    # ========================================
    # SHEET 2: DISTRIBUCIÓN DE DATOS
    # ========================================

    # Contar niveles de interés
    from collections import Counter
    interes_counter = Counter(stats['niveles_interes'])
    animo_counter = Counter(stats['estados_animo'])

    distribucion_data = {
        'CATEGORÍA': ['NIVEL DE INTERÉS'] + list(interes_counter.keys()) + [''] +
                     ['ESTADO DE ÁNIMO'] + list(animo_counter.keys()),
        'CANTIDAD': [''] + list(interes_counter.values()) + [''] +
                    [''] + list(animo_counter.values()),
        'PORCENTAJE': [''] + [f"{v/len(stats['niveles_interes'])*100:.1f}%" for v in interes_counter.values()] + [''] +
                      [''] + [f"{v/len(stats['estados_animo'])*100:.1f}%" for v in animo_counter.values()]
    }

    df_distribucion = pd.DataFrame(distribucion_data)

    # ========================================
    # SHEET 3: RECOMENDACIONES DE MEJORA
    # ========================================

    mejoras_criticas = analisis_gpt.get('mejoras_criticas', [])
    mejoras_sugeridas = analisis_gpt.get('mejoras_sugeridas', [])

    recomendaciones_data = {
        'TIPO': (['🔴 CRÍTICA'] * len(mejoras_criticas) +
                 ['🟡 SUGERIDA'] * len(mejoras_sugeridas)),
        'NÚMERO': list(range(1, len(mejoras_criticas) + 1)) +
                  list(range(len(mejoras_criticas) + 1, len(mejoras_criticas) + len(mejoras_sugeridas) + 1)),
        'RECOMENDACIÓN': mejoras_criticas + mejoras_sugeridas,
        'ESTADO': ['✅ APROBADA' if i in (mejoras_seleccionadas or []) else '⏸️ PENDIENTE'
                   for i in range(1, len(mejoras_criticas) + len(mejoras_sugeridas) + 1)]
    }

    df_recomendaciones = pd.DataFrame(recomendaciones_data)

    # ========================================
    # SHEET 4: MODIFICACIONES AL PROMPT
    # ========================================

    modificaciones = analisis_gpt.get('modificaciones_prompt', [])

    modificaciones_data = {
        'NÚMERO': list(range(1, len(modificaciones) + 1)),
        'SECCIÓN': [m['seccion'] for m in modificaciones],
        'TEXTO ORIGINAL': [m.get('texto_original', 'N/A') for m in modificaciones],
        'CAMBIO PROPUESTO (TEXTO EXACTO)': [m['cambio'] for m in modificaciones],
        'MOTIVO (CON DATOS)': [m.get('motivo', 'Mejorar desempeño basado en datos') for m in modificaciones],
        'IMPACTO ESPERADO': [m.get('impacto', 'Medio') for m in modificaciones],
        'ESTADO': ['✅ APROBADA' if i in (mejoras_seleccionadas or []) else '⏸️ PENDIENTE'
                   for i in range(1, len(modificaciones) + 1)]
    }

    df_modificaciones = pd.DataFrame(modificaciones_data)

    # ========================================
    # SHEET 5: RESUMEN ANÁLISIS GPT
    # ========================================

    resumen_gpt_data = {
        'ANÁLISIS': ['Resumen General'],
        'CONTENIDO': [analisis_gpt.get('resumen', 'Sin resumen disponible')]
    }

    df_resumen_gpt = pd.DataFrame(resumen_gpt_data)

    # ========================================
    # GUARDAR EXCEL CON FORMATO
    # ========================================

    with pd.ExcelWriter(nombre_archivo, engine='openpyxl') as writer:
        df_resumen.to_excel(writer, sheet_name='1. Resumen Ejecutivo', index=False)
        df_distribucion.to_excel(writer, sheet_name='2. Distribución', index=False)
        df_recomendaciones.to_excel(writer, sheet_name='3. Recomendaciones', index=False)
        df_modificaciones.to_excel(writer, sheet_name='4. Modificaciones Prompt', index=False)
        df_resumen_gpt.to_excel(writer, sheet_name='5. Análisis GPT', index=False)

    # ========================================
    # APLICAR FORMATO
    # ========================================

    wb = load_workbook(nombre_archivo)

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    critica_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    sugerida_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    aprobada_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Formatear cada sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Formatear headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Aplicar bordes a todas las celdas con datos
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    # Formatear sheet de recomendaciones
    if '3. Recomendaciones' in wb.sheetnames:
        ws_rec = wb['3. Recomendaciones']
        for row in range(2, ws_rec.max_row + 1):
            tipo_cell = ws_rec[f'A{row}']
            estado_cell = ws_rec[f'D{row}']

            if '🔴 CRÍTICA' in str(tipo_cell.value):
                tipo_cell.fill = critica_fill
            elif '🟡 SUGERIDA' in str(tipo_cell.value):
                tipo_cell.fill = sugerida_fill

            if '✅ APROBADA' in str(estado_cell.value):
                estado_cell.fill = aprobada_fill

    # Formatear sheet de modificaciones
    if '4. Modificaciones Prompt' in wb.sheetnames:
        ws_mod = wb['4. Modificaciones Prompt']
        for row in range(2, ws_mod.max_row + 1):
            estado_cell = ws_mod[f'G{row}']  # Ahora es columna G (antes era F)

            if '✅ APROBADA' in str(estado_cell.value):
                # Toda la fila verde si aprobada
                for col in range(1, 8):  # Ahora son 7 columnas (antes 6)
                    ws_mod.cell(row=row, column=col).fill = aprobada_fill

    wb.save(nombre_archivo)

    print(f"\n✅ Excel generado: {nombre_archivo}")
    return nombre_archivo


def mostrar_analisis_terminal(analisis):
    """
    Muestra el análisis en terminal con formato claro y números para selección

    Returns:
        tuple: (mejoras_criticas, mejoras_sugeridas, modificaciones)
    """
    stats = analisis['stats']
    analisis_gpt = analisis['analisis']

    print("\n" + "=" * 80)
    print(" " * 20 + "📊 ANÁLISIS SEMANAL - BRUCE W")
    print("=" * 80)

    print("\n🔢 MÉTRICAS PRINCIPALES:")
    print("-" * 80)
    print(f"  📞 Total de llamadas:          {stats['total_llamadas']}")
    print(f"  ✅ Aprobadas:                  {stats['aprobados']}")
    print(f"  ❌ Negadas:                    {stats['total_llamadas'] - stats['aprobados']}")
    print(f"  📈 Tasa de conversión:         {stats['tasa_conversion']:.2f}%")
    print(f"  💬 WhatsApps capturados:       {stats['whatsapps_capturados']}")
    print(f"  ⭐ Nivel de interés promedio:  {stats['interes_promedio']}")
    print(f"  😊 Estado de ánimo predominante: {stats['animo_predominante']}")

    print("\n📊 DISTRIBUCIÓN DE INTERÉS:")
    print("-" * 80)
    from collections import Counter
    interes_counter = Counter(stats['niveles_interes'])
    for nivel, cantidad in interes_counter.items():
        porcentaje = (cantidad / len(stats['niveles_interes']) * 100) if stats['niveles_interes'] else 0
        barra = "█" * int(porcentaje / 2)
        print(f"  {nivel:10} {cantidad:3} llamadas ({porcentaje:5.1f}%) {barra}")

    print("\n😊 DISTRIBUCIÓN DE ESTADO DE ÁNIMO:")
    print("-" * 80)
    animo_counter = Counter(stats['estados_animo'])
    for estado, cantidad in animo_counter.items():
        porcentaje = (cantidad / len(stats['estados_animo']) * 100) if stats['estados_animo'] else 0
        barra = "█" * int(porcentaje / 2)
        print(f"  {estado:10} {cantidad:3} llamadas ({porcentaje:5.1f}%) {barra}")

    print("\n" + "=" * 80)
    print("📋 RESUMEN DEL ANÁLISIS GPT:")
    print("=" * 80)
    print(f"\n{analisis_gpt.get('resumen', 'Sin resumen')}")

    # MEJORAS CRÍTICAS
    mejoras_criticas = analisis_gpt.get('mejoras_criticas', [])
    print("\n" + "=" * 80)
    print("🔴 MEJORAS CRÍTICAS (Alta Prioridad):")
    print("=" * 80)

    if mejoras_criticas:
        for i, mejora in enumerate(mejoras_criticas, 1):
            print(f"\n  [{i}] {mejora}")
    else:
        print("\n  ✅ No hay mejoras críticas identificadas")

    # MEJORAS SUGERIDAS
    mejoras_sugeridas = analisis_gpt.get('mejoras_sugeridas', [])
    print("\n" + "=" * 80)
    print("🟡 MEJORAS SUGERIDAS (Prioridad Media):")
    print("=" * 80)

    inicio_sugeridas = len(mejoras_criticas) + 1

    if mejoras_sugeridas:
        for i, mejora in enumerate(mejoras_sugeridas, inicio_sugeridas):
            print(f"\n  [{i}] {mejora}")
    else:
        print("\n  ℹ️ No hay mejoras sugeridas")

    # MODIFICACIONES AL PROMPT
    modificaciones = analisis_gpt.get('modificaciones_prompt', [])
    print("\n" + "=" * 80)
    print("🔧 MODIFICACIONES PROPUESTAS AL SYSTEM_PROMPT:")
    print("=" * 80)

    if modificaciones:
        for i, mod in enumerate(modificaciones, 1):
            print(f"\n  [{i}] Sección: {mod['seccion']}")
            if 'texto_original' in mod and mod['texto_original'] != 'N/A':
                print(f"      📝 Texto Original: \"{mod['texto_original'][:80]}...\"")
            print(f"      ✨ Cambio Propuesto: \"{mod['cambio'][:100]}...\"")
            if 'motivo' in mod:
                print(f"      📊 Motivo: {mod['motivo']}")
            if 'impacto' in mod:
                print(f"      🎯 Impacto: {mod['impacto']}")
    else:
        print("\n  ℹ️ No hay modificaciones propuestas")

    print("\n" + "=" * 80)

    return mejoras_criticas, mejoras_sugeridas, modificaciones


def solicitar_autorizacion():
    """
    Solicita autorización al usuario y permite seleccionar mejoras

    Returns:
        tuple: (autorizado, mejoras_seleccionadas)
    """
    print("\n" + "=" * 80)
    print(" " * 25 + "⚠️ SOLICITUD DE AUTORIZACIÓN")
    print("=" * 80)
    print("\nPara autorizar la aplicación de mejoras:")
    print("  1. Revisa el análisis anterior")
    print("  2. Ingresa los números de las mejoras que deseas aplicar (separados por coma)")
    print("     Ejemplo: 1,3,5")
    print("  3. Para aplicar TODAS las mejoras, escribe: TODAS")
    print("  4. Para CANCELAR, escribe: CANCELAR")
    print("\n  Escribe 'AUTORIZACION' seguido de los números para confirmar")
    print("  Formato: AUTORIZACION 1,3,5")
    print("=" * 80)

    respuesta = input("\n👤 Tu respuesta: ").strip().upper()

    if respuesta == "CANCELAR":
        print("\n❌ Proceso cancelado por el usuario")
        return False, []

    if not respuesta.startswith("AUTORIZACION"):
        print("\n❌ Autorización NO proporcionada. Proceso cancelado.")
        return False, []

    # Parsear números
    partes = respuesta.replace("AUTORIZACION", "").strip()

    if partes == "TODAS":
        print("\n✅ Autorización recibida: TODAS las mejoras")
        return True, "TODAS"

    try:
        mejoras_seleccionadas = [int(n.strip()) for n in partes.split(",")]
        print(f"\n✅ Autorización recibida para mejoras: {mejoras_seleccionadas}")
        return True, mejoras_seleccionadas
    except:
        print("\n❌ Formato inválido. Proceso cancelado.")
        return False, []


def ejecutar_analisis_programado():
    """
    Función principal que se ejecuta cada viernes a las 9:00 AM
    """
    print("\n\n" + "🔔" * 40)
    print(" " * 30 + "⏰ ANÁLISIS SEMANAL PROGRAMADO")
    print(" " * 25 + f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    print("🔔" * 40 + "\n")

    # Inicializar sistema de auto-mejora
    auto_mejora = AutoMejoraBruce()

    # 1. ANALIZAR SEMANA
    print("\n📊 Analizando llamadas de los últimos 7 días...")
    analisis = auto_mejora.analizar_semana()

    if not analisis:
        print("\n❌ No hay datos suficientes para analizar")
        return

    # 2. MOSTRAR EN TERMINAL
    mejoras_criticas, mejoras_sugeridas, modificaciones = mostrar_analisis_terminal(analisis)

    # 3. GENERAR EXCEL PRELIMINAR (sin aprobaciones)
    print("\n📄 Generando reporte Excel preliminar...")
    archivo_excel = generar_excel_analisis(analisis)

    # 4. SOLICITAR AUTORIZACIÓN
    autorizado, mejoras_seleccionadas = solicitar_autorizacion()

    if not autorizado:
        # Guardar historial sin aplicar cambios
        auto_mejora.guardar_historial({
            **analisis,
            'autorizado': False,
            'mejoras_seleccionadas': []
        })
        print("\n💾 Análisis guardado en historial (sin aplicar cambios)")
        return

    # 5. DETERMINAR QUÉ MEJORAS APLICAR
    total_mejoras = len(mejoras_criticas) + len(mejoras_sugeridas)

    if mejoras_seleccionadas == "TODAS":
        mejoras_aplicar = list(range(1, total_mejoras + 1))
    else:
        mejoras_aplicar = mejoras_seleccionadas

    # 6. GENERAR EXCEL FINAL (con aprobaciones marcadas)
    print("\n📄 Generando reporte Excel final con cambios aprobados...")
    archivo_excel_final = generar_excel_analisis(analisis, mejoras_aplicar)

    # 7. APLICAR MEJORAS
    print("\n🔧 Aplicando mejoras autorizadas...")

    # Filtrar modificaciones según selección
    modificaciones_aprobadas = []

    for num in mejoras_aplicar:
        if num <= len(modificaciones):
            modificaciones_aprobadas.append(modificaciones[num - 1])

    if modificaciones_aprobadas:
        print(f"\n✅ Se aplicarán {len(modificaciones_aprobadas)} modificaciones al SYSTEM_PROMPT:")
        for i, mod in enumerate(modificaciones_aprobadas, 1):
            print(f"  {i}. {mod['seccion']}: {mod['cambio'][:60]}...")

    # Guardar en historial
    analisis_final = {
        **analisis,
        'autorizado': True,
        'mejoras_seleccionadas': mejoras_aplicar,
        'modificaciones_aplicadas': modificaciones_aprobadas,
        'archivo_excel': archivo_excel_final
    }

    auto_mejora.guardar_historial(analisis_final)

    print("\n" + "=" * 80)
    print("✅ PROCESO COMPLETADO")
    print("=" * 80)
    print(f"\n📊 Reporte Excel: {archivo_excel_final}")
    print(f"💾 Historial actualizado: historial_mejoras_bruce.json")
    print(f"🔧 Modificaciones aprobadas: {len(modificaciones_aprobadas)}")
    print("\n⚠️ IMPORTANTE: Las modificaciones están documentadas en el historial.")
    print("   Para aplicarlas al código, revisa el archivo Excel y actualiza")
    print("   manualmente el SYSTEM_PROMPT en agente_ventas.py")
    print("\n" + "=" * 80)


def programar_analisis_semanal():
    """
    Programa el análisis para ejecutarse cada viernes a las 9:00 AM
    """
    print("\n" + "=" * 80)
    print("🤖 SISTEMA DE AUTO-MEJORA PROGRAMADO - BRUCE W")
    print("=" * 80)
    print("\n📅 Configuración:")
    print("  • Frecuencia: Cada viernes")
    print("  • Hora: 9:00 AM")
    print("  • Acción: Analizar llamadas de últimos 7 días")
    print("  • Requiere: Autorización manual escribiendo 'AUTORIZACION'")
    print("\n✅ Sistema programado y en espera...")
    print("=" * 80 + "\n")

    # Programar tarea
    schedule.every().friday.at("09:00").do(ejecutar_analisis_programado)

    # Loop infinito para mantener el scheduler activo
    print("⏳ Esperando próximo viernes a las 9:00 AM...")
    print("   (Presiona Ctrl+C para detener)\n")

    try:
        while True:
            schedule.run_pending()
            time.sleep(60)  # Verificar cada minuto
    except KeyboardInterrupt:
        print("\n\n⛔ Sistema de auto-mejora detenido por el usuario")
        print("=" * 80 + "\n")


if __name__ == "__main__":
    import sys
    import io

    # Configurar encoding UTF-8 para Windows
    if sys.platform == 'win32':
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        # Modo testing: ejecutar inmediatamente
        print("\n🧪 MODO TESTING - Ejecutando análisis inmediato\n")
        ejecutar_analisis_programado()
    else:
        # Modo normal: programar para viernes 9:00 AM
        programar_analisis_semanal()
