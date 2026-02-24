"""
Generador de reportes Excel por lote de auditoría.

Crea un Excel con múltiples hojas:
  - Resumen: KPIs por lote (bugs, calificación, conversión)
  - Bugs: Detalle de bugs por BRUCE ID
  - Tendencia: Evolución lote a lote
  - Logs Procesados: Tracking de qué logs ya se revisaron

Uso:
  python scripts/batch_report_excel.py                    # Genera/actualiza reporte
  python scripts/batch_report_excel.py --lote 3           # Ver detalle de lote 3
  python scripts/batch_report_excel.py --path reporte.xlsx # Ruta personalizada
"""

import os
import sys
import json
from datetime import datetime

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
AUDIT_DATA_DIR = os.path.join(PROJECT_DIR, 'audit_data')
DEFAULT_EXCEL_PATH = os.path.join(AUDIT_DATA_DIR, 'auditoria_bruce.xlsx')
BATCH_HISTORY_PATH = os.path.join(AUDIT_DATA_DIR, 'batch_history.json')

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ============================================================
# COLORES Y ESTILOS
# ============================================================

_HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid') if HAS_OPENPYXL else None
_HEADER_FONT = Font(color='FFFFFF', bold=True, size=11) if HAS_OPENPYXL else None
_GOOD_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid') if HAS_OPENPYXL else None
_WARN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') if HAS_OPENPYXL else None
_BAD_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid') if HAS_OPENPYXL else None
_TITLE_FONT = Font(bold=True, size=14) if HAS_OPENPYXL else None
_SUBTITLE_FONT = Font(bold=True, size=11) if HAS_OPENPYXL else None
_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
) if HAS_OPENPYXL else None


def _style_header_row(ws, row, max_col):
    """Aplica estilos a fila de encabezado."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = _BORDER


def _auto_width(ws):
    """Ajusta ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


# ============================================================
# BATCH HISTORY MANAGEMENT
# ============================================================

def load_batch_history():
    """Carga historial de lotes."""
    os.makedirs(AUDIT_DATA_DIR, exist_ok=True)
    if os.path.exists(BATCH_HISTORY_PATH):
        with open(BATCH_HISTORY_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'batches': [], 'logs_processed': {}}


def save_batch_history(history):
    """Guarda historial de lotes."""
    os.makedirs(AUDIT_DATA_DIR, exist_ok=True)
    with open(BATCH_HISTORY_PATH, 'w', encoding='utf-8') as f:
        json.dump(history, f, ensure_ascii=False, indent=2)


def add_batch(audit_result, history=None):
    """Agrega resultado de auditoría como nuevo lote.

    Args:
        audit_result: dict con métricas del lote
        history: historial existente (se carga si None)

    Returns:
        número de lote asignado
    """
    if history is None:
        history = load_batch_history()

    batch_num = len(history['batches']) + 1
    batch = {
        'lote': batch_num,
        'timestamp': datetime.now().isoformat(),
        'llamadas': audit_result.get('total_llamadas', 0),
        'bruce_ids': audit_result.get('bruce_ids', []),
        'bugs_total': audit_result.get('bugs_total', 0),
        'bugs_criticos': audit_result.get('bugs_criticos', 0),
        'bugs_por_tipo': audit_result.get('bugs_por_tipo', {}),
        'calificacion_promedio': audit_result.get('calificacion_promedio', 0),
        'conversion_rate': audit_result.get('conversion_rate', 0),
        'regression_pass_rate': audit_result.get('regression_pass_rate', 0),
        'logs_procesados': audit_result.get('logs_procesados', []),
        'version_deploy': audit_result.get('version_deploy', ''),
        'fix_aplicados': audit_result.get('fix_aplicados', []),
        'notas': audit_result.get('notas', ''),
        # Métricas detalladas
        'bugs_detalle': audit_result.get('bugs_detalle', []),
        'pattern_audit': audit_result.get('pattern_audit', {}),
        'endpoints_ok': audit_result.get('endpoints_ok', 0),
        'endpoints_total': audit_result.get('endpoints_total', 0),
    }

    history['batches'].append(batch)

    # Marcar logs como procesados
    for log_file in audit_result.get('logs_procesados', []):
        history['logs_processed'][log_file] = {
            'lote': batch_num,
            'fecha': datetime.now().isoformat(),
        }

    save_batch_history(history)
    return batch_num


# ============================================================
# GENERACIÓN EXCEL
# ============================================================

def generate_excel(history=None, output_path=None):
    """Genera reporte Excel completo desde el historial de lotes.

    Args:
        history: historial de lotes (se carga si None)
        output_path: ruta del Excel (default: audit_data/auditoria_bruce.xlsx)

    Returns:
        ruta del archivo generado
    """
    if not HAS_OPENPYXL:
        print("[!] openpyxl no disponible. Instalando...")
        os.system(f'{sys.executable} -m pip install openpyxl -q')
        print("[!] Reinicia el script después de instalar openpyxl")
        return None

    if history is None:
        history = load_batch_history()

    if output_path is None:
        output_path = DEFAULT_EXCEL_PATH

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()

    # -------- Hoja 1: RESUMEN --------
    ws_resumen = wb.active
    ws_resumen.title = "Resumen por Lote"
    _write_resumen(ws_resumen, history)

    # -------- Hoja 2: BUGS DETALLE --------
    ws_bugs = wb.create_sheet("Bugs Detalle")
    _write_bugs_detalle(ws_bugs, history)

    # -------- Hoja 3: TENDENCIA --------
    ws_tendencia = wb.create_sheet("Tendencia")
    _write_tendencia(ws_tendencia, history)

    # -------- Hoja 4: LOGS PROCESADOS --------
    ws_logs = wb.create_sheet("Logs Procesados")
    _write_logs_procesados(ws_logs, history)

    wb.save(output_path)
    return output_path


def _write_resumen(ws, history):
    """Hoja de resumen KPI por lote."""
    # Título
    ws['A1'] = 'AUDITORÍA BRUCE W - RESUMEN POR LOTE'
    ws['A1'].font = _TITLE_FONT
    ws.merge_cells('A1:L1')

    ws['A2'] = f'Generado: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].font = Font(italic=True, color='666666')

    # Headers
    row = 4
    headers = [
        'Lote', 'Fecha', 'Llamadas', 'Bruce IDs',
        'Bugs Total', 'Bugs Críticos', 'Bug Rate',
        'Calif. Prom.', 'Conversión %', 'Regresión %',
        'Deploy', 'FIX Aplicados'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    # Datos
    batches = history.get('batches', [])
    for i, batch in enumerate(batches):
        r = row + 1 + i
        llamadas = batch.get('llamadas', 0)
        bugs = batch.get('bugs_total', 0)
        bug_rate = f"{(bugs/llamadas*100):.1f}%" if llamadas > 0 else "N/A"

        ws.cell(row=r, column=1, value=batch.get('lote', i+1))
        ws.cell(row=r, column=2, value=batch.get('timestamp', '')[:10])
        ws.cell(row=r, column=3, value=llamadas)
        ws.cell(row=r, column=4, value=len(batch.get('bruce_ids', [])))
        ws.cell(row=r, column=5, value=bugs)
        ws.cell(row=r, column=6, value=batch.get('bugs_criticos', 0))
        ws.cell(row=r, column=7, value=bug_rate)
        ws.cell(row=r, column=8, value=batch.get('calificacion_promedio', 0))
        ws.cell(row=r, column=9, value=f"{batch.get('conversion_rate', 0):.1f}%")
        ws.cell(row=r, column=10, value=f"{batch.get('regression_pass_rate', 0):.1f}%")
        ws.cell(row=r, column=11, value=batch.get('version_deploy', '')[:15])
        ws.cell(row=r, column=12, value=', '.join(batch.get('fix_aplicados', []))[:50])

        # Color por bugs
        bug_cell = ws.cell(row=r, column=5)
        crit_cell = ws.cell(row=r, column=6)
        if bugs == 0:
            bug_cell.fill = _GOOD_FILL
        elif bugs <= 3:
            bug_cell.fill = _WARN_FILL
        else:
            bug_cell.fill = _BAD_FILL

        if batch.get('bugs_criticos', 0) > 0:
            crit_cell.fill = _BAD_FILL

        # Color por calificación
        cal_cell = ws.cell(row=r, column=8)
        cal = batch.get('calificacion_promedio', 0)
        if cal >= 8:
            cal_cell.fill = _GOOD_FILL
        elif cal >= 6:
            cal_cell.fill = _WARN_FILL
        elif cal > 0:
            cal_cell.fill = _BAD_FILL

        # Bordes
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = _BORDER
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')

    # Resumen global
    if batches:
        summary_row = row + len(batches) + 2
        ws.cell(row=summary_row, column=1, value='TOTALES').font = _SUBTITLE_FONT
        total_llamadas = sum(b.get('llamadas', 0) for b in batches)
        total_bugs = sum(b.get('bugs_total', 0) for b in batches)
        ws.cell(row=summary_row, column=3, value=total_llamadas)
        ws.cell(row=summary_row, column=5, value=total_bugs)
        ws.cell(row=summary_row, column=7,
                value=f"{(total_bugs/total_llamadas*100):.1f}%" if total_llamadas > 0 else "N/A")

        # Tendencia: último vs anterior
        if len(batches) >= 2:
            prev = batches[-2]
            curr = batches[-1]
            trend_row = summary_row + 1
            ws.cell(row=trend_row, column=1, value='TENDENCIA').font = _SUBTITLE_FONT

            prev_rate = prev.get('bugs_total', 0) / max(prev.get('llamadas', 1), 1)
            curr_rate = curr.get('bugs_total', 0) / max(curr.get('llamadas', 1), 1)
            change = curr_rate - prev_rate
            ws.cell(row=trend_row, column=7,
                    value=f"{'↑' if change > 0 else '↓'} {abs(change)*100:.1f}%")

    _auto_width(ws)


def _write_bugs_detalle(ws, history):
    """Hoja con detalle de bugs por lote."""
    ws['A1'] = 'DETALLE DE BUGS POR LOTE'
    ws['A1'].font = _TITLE_FONT

    row = 3
    headers = ['Lote', 'Bruce ID', 'Tipo Bug', 'Severidad', 'Detalle']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    r = row + 1
    for batch in history.get('batches', []):
        lote = batch.get('lote', '?')
        for bug in batch.get('bugs_detalle', []):
            ws.cell(row=r, column=1, value=lote)
            ws.cell(row=r, column=2, value=bug.get('bruce_id', ''))
            ws.cell(row=r, column=3, value=bug.get('tipo', ''))
            ws.cell(row=r, column=4, value=bug.get('severidad', ''))
            ws.cell(row=r, column=5, value=bug.get('detalle', '')[:100])

            # Color por severidad
            sev = bug.get('severidad', '')
            sev_cell = ws.cell(row=r, column=4)
            if sev == 'CRITICO':
                sev_cell.fill = _BAD_FILL
            elif sev == 'ALTO':
                sev_cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
            elif sev == 'MEDIO':
                sev_cell.fill = _WARN_FILL

            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).border = _BORDER

            r += 1

        # Si tiene bugs por tipo (agregado)
        if not batch.get('bugs_detalle') and batch.get('bugs_por_tipo'):
            for tipo, count in batch['bugs_por_tipo'].items():
                ws.cell(row=r, column=1, value=lote)
                ws.cell(row=r, column=3, value=tipo)
                ws.cell(row=r, column=5, value=f"{count} ocurrencia(s)")
                for col in range(1, len(headers) + 1):
                    ws.cell(row=r, column=col).border = _BORDER
                r += 1

    _auto_width(ws)


def _write_tendencia(ws, history):
    """Hoja de tendencia con gráfica."""
    ws['A1'] = 'TENDENCIA DE CALIDAD POR LOTE'
    ws['A1'].font = _TITLE_FONT

    batches = history.get('batches', [])
    if not batches:
        ws['A3'] = 'Sin datos de lotes aún'
        return

    # Tabla de datos
    row = 3
    headers = ['Lote', 'Bugs Total', 'Bug Rate %', 'Calif.', 'Conversión %', 'Regresión %']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    for i, batch in enumerate(batches):
        r = row + 1 + i
        llamadas = max(batch.get('llamadas', 1), 1)
        ws.cell(row=r, column=1, value=f"Lote {batch.get('lote', i+1)}")
        ws.cell(row=r, column=2, value=batch.get('bugs_total', 0))
        ws.cell(row=r, column=3, value=round(batch.get('bugs_total', 0) / llamadas * 100, 1))
        ws.cell(row=r, column=4, value=batch.get('calificacion_promedio', 0))
        ws.cell(row=r, column=5, value=round(batch.get('conversion_rate', 0), 1))
        ws.cell(row=r, column=6, value=round(batch.get('regression_pass_rate', 0), 1))

    # Gráfica de barras - Bugs por lote
    if len(batches) >= 2:
        chart = BarChart()
        chart.type = 'col'
        chart.title = 'Bugs por Lote'
        chart.y_axis.title = 'Cantidad'
        chart.x_axis.title = 'Lote'

        data_ref = Reference(ws, min_col=2, min_row=row, max_row=row + len(batches), max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=row + 1, max_row=row + len(batches))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        ws.add_chart(chart, 'H3')

        # Gráfica de línea - Calificación
        line_chart = LineChart()
        line_chart.title = 'Calificación Promedio'
        line_chart.y_axis.title = 'Calificación'
        line_chart.y_axis.scaling.min = 0
        line_chart.y_axis.scaling.max = 10

        data_ref2 = Reference(ws, min_col=4, min_row=row, max_row=row + len(batches), max_col=4)
        line_chart.add_data(data_ref2, titles_from_data=True)
        line_chart.set_categories(cats_ref)
        ws.add_chart(line_chart, 'H20')

    _auto_width(ws)


def _write_logs_procesados(ws, history):
    """Hoja de tracking de logs procesados."""
    ws['A1'] = 'LOGS PROCESADOS'
    ws['A1'].font = _TITLE_FONT

    ws['A2'] = 'Archivos de log ya revisados (no se vuelven a procesar)'
    ws['A2'].font = Font(italic=True, color='666666')

    row = 4
    headers = ['Archivo', 'Lote', 'Fecha Procesamiento']
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header_row(ws, row, len(headers))

    logs_processed = history.get('logs_processed', {})
    for i, (log_file, info) in enumerate(sorted(logs_processed.items())):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=log_file)
        ws.cell(row=r, column=2, value=info.get('lote', '?'))
        ws.cell(row=r, column=3, value=info.get('fecha', '')[:19])
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = _BORDER

    _auto_width(ws)


# ============================================================
# CLI
# ============================================================

def main():
    import argparse
    parser = argparse.ArgumentParser(description='Bruce W - Batch Report Excel')
    parser.add_argument('--path', type=str, default=DEFAULT_EXCEL_PATH,
                        help='Ruta del Excel')
    parser.add_argument('--lote', type=int, help='Ver detalle de un lote específico')
    parser.add_argument('--add-test', action='store_true',
                        help='Agregar un lote de prueba')
    args = parser.parse_args()

    history = load_batch_history()

    if args.add_test:
        # Lote de prueba
        test_batch = {
            'total_llamadas': 25,
            'bruce_ids': [f'BRUCE{2500+i}' for i in range(25)],
            'bugs_total': 3,
            'bugs_criticos': 0,
            'bugs_por_tipo': {'PREGUNTA_REPETIDA': 1, 'PITCH_REPETIDO': 1, 'CLIENTE_HABLA_ULTIMO': 1},
            'calificacion_promedio': 7.5,
            'conversion_rate': 12.0,
            'regression_pass_rate': 98.5,
            'logs_procesados': ['test_log.txt'],
            'version_deploy': 'test-v1',
            'fix_aplicados': ['FIX 798'],
            'bugs_detalle': [
                {'bruce_id': 'BRUCE2501', 'tipo': 'PREGUNTA_REPETIDA', 'severidad': 'MEDIO',
                 'detalle': 'Pidió WhatsApp 2 veces'},
            ],
        }
        batch_num = add_batch(test_batch, history)
        print(f"Lote de prueba #{batch_num} agregado")
        history = load_batch_history()

    if args.lote:
        batches = history.get('batches', [])
        match = [b for b in batches if b.get('lote') == args.lote]
        if match:
            print(json.dumps(match[0], ensure_ascii=False, indent=2))
        else:
            print(f"Lote {args.lote} no encontrado")
        return

    # Generar Excel
    print(f"\n{'='*50}")
    print(f"  GENERANDO REPORTE EXCEL")
    print(f"{'='*50}\n")

    path = generate_excel(history, args.path)
    if path:
        total_lotes = len(history.get('batches', []))
        total_logs = len(history.get('logs_processed', {}))
        print(f"  Lotes registrados: {total_lotes}")
        print(f"  Logs procesados: {total_logs}")
        print(f"  Excel generado: {path}")
    else:
        print("  Error generando Excel")


if __name__ == '__main__':
    main()
